import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import holidays
from streamlit_calendar import calendar
import json
import os
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. CONFIGURATION ET CONSTANTES
# ==========================================
st.set_page_config(page_title="Planning RI Pro", layout="wide", initial_sidebar_state="expanded")

JOURS_FR = {"Monday": "Lundi", "Tuesday": "Mardi", "Wednesday": "Mercredi",
            "Thursday": "Jeudi", "Friday": "Vendredi", "Saturday": "Samedi", "Sunday": "Dimanche"}
MOIS_FR = {1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril", 5: "Mai", 6: "Juin",
           7: "Juillet", 8: "Août", 9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"}

FICHIER_SAUVEGARDE = "equipe_ri.json"

# ==========================================
# 2. THÈME VISUEL (CSS - DYNAMIQUE CLAIR/SOMBRE)
# ==========================================
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;800&display=swap');
    
    :root {
        --bg-main: #F0F9FF; --bg-sec: #E0F2FE; --bg-card: #FFFFFF;
        --text-main: #333333; --text-title: #000000;
        --border-color: #BAE6FD; --hover-bg: #DBEAFE;
    }
    @media (prefers-color-scheme: dark) {
        :root {
            --bg-main: #0F172A; --bg-sec: #1E293B; --bg-card: #334155;
            --text-main: #F8FAFC; --text-title: #FFFFFF;
            --border-color: #475569; --hover-bg: #475569;
        }
    }

    html, body, [class*="css"], .stApp, [data-testid="stSidebar"] { font-family: 'Inter', sans-serif; }
    p, span, label, div, th, td, input { color: var(--text-main) !important; }
    h1, h2, h3 { color: var(--text-title) !important; }
    svg { fill: var(--text-main) !important; }
    
    .stApp, header[data-testid="stHeader"], div[role="dialog"] { background-color: var(--bg-main) !important; } 
    [data-testid="stSidebar"] { background-color: var(--bg-sec) !important; } 
    
    @media (prefers-color-scheme: dark) { [data-testid="stModal"] > div:first-child { background-color: rgba(0, 0, 0, 0.7) !important; } }
    @media (prefers-color-scheme: light) { [data-testid="stModal"] > div:first-child { background-color: rgba(3, 105, 161, 0.5) !important; } }
    
    div[data-baseweb="input"] > div, [data-testid="stExpander"] { 
        background-color: var(--bg-card) !important; border: 1px solid var(--border-color) !important; border-radius: 8px; 
    }
    
    div[data-baseweb="popover"], div[data-baseweb="popover"] > div, div[data-baseweb="calendar"],
    div[data-baseweb="calendar"] > div, div[data-baseweb="calendar"] > div > div,
    div[data-baseweb="calendar"] [role="heading"], div[data-baseweb="calendar"] [role="grid"],
    div[data-baseweb="calendar"] [role="row"], div[data-baseweb="calendar"] [role="rowheader"],
    div[data-baseweb="calendar"] [role="columnheader"] {
        background-color: var(--bg-card) !important;
    }
    div[data-baseweb="calendar"] [role="button"] { background-color: transparent !important; color: var(--text-main) !important; }
    div[data-baseweb="calendar"] [role="button"]:hover { background-color: var(--hover-bg) !important; }
    div[data-baseweb="calendar"] [aria-selected="true"], div[data-baseweb="calendar"] [aria-selected="true"] * {
        background-color: #2563EB !important; color: #FFFFFF !important;
    }

    div[role="dialog"] { border: 2px solid var(--border-color) !important; border-radius: 12px; }
    h1 { border-bottom: 4px solid #2563EB; padding-bottom: 10px; margin-bottom: 1rem; }
    
    .stButton>button { border-radius: 8px; border: 1px solid var(--border-color) !important; background-color: var(--bg-card) !important; transition: all 0.2s; font-weight: 600; }
    .stButton>button:hover { border-color: #2563EB !important; background-color: var(--hover-bg) !important; transform: translateY(-1px); }
    .stButton>button p { color: var(--text-main) !important; }
    
    .btn-valider button *, .btn-generer button *, .btn-supprimer button *, .btn-indispo button *, .btn-obli1 button *, .btn-obli2 button *, .btn-clear button * { color: white !important; fill: white !important; }
    .btn-valider button { background-color: #059669 !important; border: none !important; }
    .btn-generer button { background: linear-gradient(135deg, #2563EB 0%, #1D4ED8 100%) !important; padding: 1rem !important; border: none !important;}
    .btn-supprimer button { background-color: #DC2626 !important; padding: 2px 10px !important; border: none !important;}
    
    /* Boutons de la modale */
    .btn-indispo button { background-color: #DC2626 !important; border: none !important;}
    .btn-obli1 button { background-color: #16A34A !important; border: none !important;} /* Vert */
    .btn-obli2 button { background-color: #F59E0B !important; border: none !important;} /* Orange */
    .btn-clear button { background-color: #64748B !important; border: none !important;} /* Gris ardoise */

    .stTabs [data-baseweb="tab-list"] { background-color: var(--bg-sec) !important; padding: 6px; border-radius: 10px; }
    .stTabs [aria-selected="true"] { background-color: var(--bg-card) !important; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
    .stTabs [aria-selected="true"] p { color: var(--text-title) !important; font-weight: bold; }
    [data-testid="stTable"] { background-color: var(--bg-card) !important; border-radius: 8px; border: 1px solid var(--border-color); }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 3. GESTION DES DONNÉES (JSON)
# ==========================================
def charger_donnees():
    if os.path.exists(FICHIER_SAUVEGARDE):
        with open(FICHIER_SAUVEGARDE, "r", encoding="utf-8") as f:
            data = json.load(f)
            for m in data.values():
                m.setdefault("score_we", 0)
                m.setdefault("nb_l1", 0)
                m.setdefault("nb_l2", 0)
                m.setdefault("lignes", [1, 2])
                m.setdefault("obl_l1", []) # NOUVEAU: Liste bonus L1
                m.setdefault("obl_l2", []) # NOUVEAU: Liste bonus L2
            return data
    return None

def sauvegarder_donnees(data):
    with open(FICHIER_SAUVEGARDE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)

def initialiser_etat():
    if 'merms_data' not in st.session_state:
        donnees = charger_donnees()
        if not donnees:
            noms = ["Lechevin L.", "Abdelaoui F.", "Laurin M.", "Cotton L.", "Bacquet V.", 
                    "Leroux C.", "Brasseur O.", "Dupierris P.A.", "Talbaut V.", "Michel L.", "Dhondt F.", "Geffroy C."]
            donnees = {
                name: {"lignes": [2] if name in ["Dhondt F.", "Geffroy C."] else [1, 2],
                       "score_cumule": 0, "score_we": 0, "nb_l1": 0, "nb_l2": 0,
                       "pref_vendredi": False, "absences": [], "obl_l1": [], "obl_l2": []} for name in noms
            }
            sauvegarder_donnees(donnees)
        st.session_state.merms_data = donnees
    
    if 'modal_ouvert' not in st.session_state:
        st.session_state.modal_ouvert = None

initialiser_etat()

# ==========================================
# 4. COMPOSANT CALENDRIER (DESIDERATA)
# ==========================================
@st.dialog("Saisie des Desiderata", width="large")
def modal_desiderata(name):
    st.write(f"### 👤 Agent : **{name}**")
    st.info("📱 **Sur smartphone :** Tapotez simplement un jour pour le sélectionner, ou glissez pour sélectionner une semaine.")
    
    t_abs, t_o1, t_o2, t_sel, l_sel = f"abs_{name}", f"o1_{name}", f"o2_{name}", f"sel_{name}", f"lsel_{name}"
    
    if t_abs not in st.session_state: st.session_state[t_abs] = st.session_state.merms_data[name]["absences"].copy()
    if t_o1 not in st.session_state: st.session_state[t_o1] = st.session_state.merms_data[name].get("obl_l1", []).copy()
    if t_o2 not in st.session_state: st.session_state[t_o2] = st.session_state.merms_data[name].get("obl_l2", []).copy()
    if t_sel not in st.session_state: st.session_state[t_sel] = set() 

    deb_str, fin_str = st.session_state.d_start.strftime("%Y-%m-%d"), (st.session_state.d_end + timedelta(days=1)).strftime("%Y-%m-%d")

    cal_options = {
        "timeZone": "UTC", "initialDate": deb_str, "validRange": {"start": deb_str, "end": fin_str},
        "headerToolbar": {"left": "prev,next", "center": "title", "right": ""},
        "selectable": True, "locale": "fr", "firstDay": 1, "height": "450px", "unselectAuto": False,
        "selectLongPressDelay": 50, "longPressDelay": 50
    }
    
    css_cal = """
        :root {
            --fc-bg: #F0F9FF; --fc-text: #333333; --fc-border: #BAE6FD;
            --fc-header: #E0F2FE; --fc-today: #DBEAFE; --fc-other: #FFFFFF;
        }
        @media (prefers-color-scheme: dark) {
            :root {
                --fc-bg: #0F172A; --fc-text: #F8FAFC; --fc-border: #475569;
                --fc-header: #1E293B; --fc-today: #334155; --fc-other: #1E293B;
            }
        }
        body { background-color: var(--fc-bg) !important; }
        .fc { background-color: var(--fc-bg) !important; color: var(--fc-text) !important; font-family: 'Inter', sans-serif; }
        .fc-theme-standard th, .fc-theme-standard td { border-color: var(--fc-border) !important; }
        .fc-col-header-cell { background-color: var(--fc-header) !important; }
        .fc-button { background-color: #0284C7 !important; border-color: #0284C7 !important; color: white !important; background-image: none !important; box-shadow: none !important; text-transform: capitalize; }
        .fc-button:hover { background-color: #0369A1 !important; border-color: #0369A1 !important; }
        .fc-button:disabled { background-color: #7DD3FC !important; border-color: #7DD3FC !important; color: white !important; opacity: 1 !important; }
        .fc-day-other { background-color: var(--fc-other) !important; opacity: 0.6; }
        .fc-day-today { background-color: var(--fc-today) !important; }
        .fc-highlight { background-color: rgba(37, 99, 235, 0.4) !important; } 
        .fc-toolbar-title, .fc-daygrid-day-number { color: var(--fc-text) !important; }
    """
    
    # Affichage des 4 types d'événements
    events = [{"title": "INDISPO", "start": d, "end": d, "color": "#DC2626"} for d in st.session_state[t_abs] if d not in st.session_state[t_sel]] + \
             [{"title": "OBLI L1", "start": d, "end": d, "color": "#16A34A"} for d in st.session_state[t_o1] if d not in st.session_state[t_sel]] + \
             [{"title": "OBLI L2", "start": d, "end": d, "color": "#F59E0B"} for d in st.session_state[t_o2] if d not in st.session_state[t_sel]] + \
             [{"title": "SÉLECTION", "start": d, "end": d, "color": "#3B82F6"} for d in st.session_state[t_sel]]

    res = calendar(events=events, options=cal_options, custom_css=css_cal, key=f"cal_{name}")

    action_detectee = False
    if "select" in res and str(res["select"]) != st.session_state.get(l_sel):
        st.session_state[l_sel] = str(res["select"])
        s_date, e_date = res["select"].get("startStr", res["select"]["start"])[:10], res["select"].get("endStr", res["select"]["end"])[:10]
        st.session_state[t_sel].update(pd.date_range(s_date, pd.to_datetime(e_date) - pd.Timedelta(days=1)).strftime("%Y-%m-%d").tolist())
        action_detectee = True

    elif "dateClick" in res and str(res["dateClick"]) != st.session_state.get(l_sel):
        st.session_state[l_sel] = str(res["dateClick"])
        st.session_state[t_sel].add(res["dateClick"].get("dateStr", res["dateClick"]["date"])[:10])
        action_detectee = True

    if action_detectee: st.rerun()

    # DISPOSITION 2x2 POUR LES BOUTONS D'ACTION
    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="btn-indispo">', unsafe_allow_html=True)
        if st.button("🔴 Indisponible (L1/L2)", use_container_width=True):
            st.session_state[t_abs] = list(set(st.session_state[t_abs] + list(st.session_state[t_sel])))
            st.session_state[t_o1] = [d for d in st.session_state[t_o1] if d not in st.session_state[t_sel]]
            st.session_state[t_o2] = [d for d in st.session_state[t_o2] if d not in st.session_state[t_sel]]
            st.session_state[t_sel].clear(); st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    with c2:
        st.markdown('<div class="btn-obli1">', unsafe_allow_html=True)
        if st.button("🟢 Obligatoire LIGNE 1", use_container_width=True):
            st.session_state[t_o1] = list(set(st.session_state[t_o1] + list(st.session_state[t_sel])))
            st.session_state[t_abs] = [d for d in st.session_state[t_abs] if d not in st.session_state[t_sel]]
            st.session_state[t_o2] = [d for d in st.session_state[t_o2] if d not in st.session_state[t_sel]]
            st.session_state[t_sel].clear(); st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    c3, c4 = st.columns(2)
    with c3:
        st.markdown('<div class="btn-obli2">', unsafe_allow_html=True)
        if st.button("🟠 Obligatoire LIGNE 2", use_container_width=True):
            st.session_state[t_o2] = list(set(st.session_state[t_o2] + list(st.session_state[t_sel])))
            st.session_state[t_abs] = [d for d in st.session_state[t_abs] if d not in st.session_state[t_sel]]
            st.session_state[t_o1] = [d for d in st.session_state[t_o1] if d not in st.session_state[t_sel]]
            st.session_state[t_sel].clear(); st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    with c4:
        st.markdown('<div class="btn-clear">', unsafe_allow_html=True)
        if st.button("⚪ Jour Normal (Neutre)", use_container_width=True):
            st.session_state[t_abs] = [d for d in st.session_state[t_abs] if d not in st.session_state[t_sel]]
            st.session_state[t_o1] = [d for d in st.session_state[t_o1] if d not in st.session_state[t_sel]]
            st.session_state[t_o2] = [d for d in st.session_state[t_o2] if d not in st.session_state[t_sel]]
            st.session_state[t_sel].clear(); st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    st.write("---")
    v_pref = st.toggle("Coupler le vendredi au WE (lorsque je suis d'astreinte)", value=st.session_state.merms_data[name]["pref_vendredi"])
    st.write("---")
    
    st.markdown('<div class="btn-valider">', unsafe_allow_html=True)
    if st.button("✅ CONFIRMER ET ENREGISTRER MES CHOIX", use_container_width=True):
        st.session_state.merms_data[name]["absences"] = st.session_state[t_abs]
        st.session_state.merms_data[name]["obl_l1"] = st.session_state[t_o1]
        st.session_state.merms_data[name]["obl_l2"] = st.session_state[t_o2]
        st.session_state.merms_data[name]["pref_vendredi"] = v_pref
        del st.session_state[t_abs], st.session_state[t_o1], st.session_state[t_o2], st.session_state[t_sel]
        sauvegarder_donnees(st.session_state.merms_data)
        st.session_state.modal_ouvert = None
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

# ==========================================
# 5. MOTEUR ALGORITHMIQUE (ÉQUITÉ & RÈGLES)
# ==========================================
def generer_planning(debut, fin):
    debut, fin = pd.Timestamp(debut), pd.Timestamp(fin)
    fr_holidays = holidays.France(years=range(debut.year, fin.year + 1))
    jours = pd.date_range(debut, fin)
    
    planning = {d: {"L1": "⚠️ À POURVOIR", "L2": "⚠️ À POURVOIR"} for d in jours}
    assigned_dates = {m: set() for m in st.session_state.merms_data}
    
    sc, sc_we = {m: v['score_cumule'] for m, v in st.session_state.merms_data.items()}, {m: v['score_we'] for m, v in st.session_state.merms_data.items()}
    n_l1, n_l2 = {m: v['nb_l1'] for m, v in st.session_state.merms_data.items()}, {m: v['nb_l2'] for m, v in st.session_state.merms_data.items()}
    
    def est_dispo(m, dt_list): return not any(d.strftime("%Y-%m-%d") in st.session_state.merms_data[m]["absences"] for d in dt_list)
    def a_prio_l1(m, dt_list): return any(d.strftime("%Y-%m-%d") in st.session_state.merms_data[m].get("obl_l1", []) for d in dt_list)
    def a_prio_l2(m, dt_list): return any(d.strftime("%Y-%m-%d") in st.session_state.merms_data[m].get("obl_l2", []) for d in dt_list)
    def is_we_ferie(d): return d.weekday() >= 5 or d.date() in fr_holidays

    # --- PASSE 1 : WEEK-ENDS ---
    for d in jours:
        if d.weekday() == 5:
            we_days = [d, d + timedelta(days=1)] if (d + timedelta(days=1)) <= fin else [d]
            d_fri = d - timedelta(days=1)
            
            for ligne in ["L1", "L2"]:
                candidats = []
                for m, v in st.session_state.merms_data.items():
                    if int(ligne[1]) not in v["lignes"] or (ligne == "L2" and planning[d]["L1"] == m): continue
                    if not est_dispo(m, we_days): continue
                    if v["pref_vendredi"] and d_fri >= debut and (not est_dispo(m, [d_fri]) or planning[d_fri][ligne] != "⚠️ À POURVOIR"): continue 
                    candidats.append(m)
                
                if candidats:
                    # Règle de l'Astreinte Obligatoire ciblée
                    jours_concernes = we_days + ([d_fri] if d_fri >= debut else [])
                    if ligne == "L1":
                        candidats_prio = [c for c in candidats if a_prio_l1(c, jours_concernes)]
                    else:
                        candidats_prio = [c for c in candidats if a_prio_l2(c, jours_concernes)]
                        
                    if candidats_prio: candidats = candidats_prio # S'il y a des volontaires, on ne garde qu'eux
                    
                    choix = min(candidats, key=lambda x: (sc_we[x], sc[x], n_l1[x] + n_l2[x], n_l1[x] if ligne == "L1" else n_l2[x]))
                    sc_we[choix] += 1 
                    
                    jours_assign = we_days + ([d_fri] if (st.session_state.merms_data[choix]["pref_vendredi"] and d_fri >= debut) else [])
                        
                    for ja in jours_assign:
                        planning[ja][ligne] = choix
                        assigned_dates[choix].add(ja)
                        sc[choix] += 3 if is_we_ferie(ja) else 1
                        if ligne == "L1": n_l1[choix] += 1
                        else: n_l2[choix] += 1
                        
    # --- PASSE 2 : SEMAINE ---
    for d in jours:
        for ligne in ["L1", "L2"]:
            if planning[d][ligne] != "⚠️ À POURVOIR": continue
            
            candidats = []
            for m, v in st.session_state.merms_data.items():
                if int(ligne[1]) not in v["lignes"] or (ligne == "L2" and planning[d]["L1"] == m) or not est_dispo(m, [d]): continue
                if (d - timedelta(days=1)) in assigned_dates[m] or (d + timedelta(days=1)) in assigned_dates[m]: continue
                
                jours_sem = [ad for ad in assigned_dates[m] if ad.isocalendar()[1] == d.isocalendar()[1]]
                a_un_we = any(ad.weekday() >= 5 for ad in jours_sem)
                jours_hors_we = [ad for ad in jours_sem if ad.weekday() < 5]
                
                if (a_un_we and len(jours_hors_we) >= 1) or (not a_un_we and len(jours_hors_we) >= 2): continue 
                candidats.append(m)
                
            if candidats:
                # Règle de l'Astreinte Obligatoire ciblée
                if ligne == "L1":
                    candidats_prio = [c for c in candidats if a_prio_l1(c, [d])]
                else:
                    candidats_prio = [c for c in candidats if a_prio_l2(c, [d])]
                    
                if candidats_prio: candidats = candidats_prio 
                
                choix = min(candidats, key=lambda x: (sc[x], n_l1[x] + n_l2[x], n_l1[x] if ligne == "L1" else n_l2[x]))
                planning[d][ligne] = choix
                assigned_dates[choix].add(d)
                sc[choix] += 3 if is_we_ferie(d) else 1
                if ligne == "L1": n_l1[choix] += 1
                else: n_l2[choix] += 1

    res_df = pd.DataFrame([{
        "Date": d.strftime("%d/%m/%Y"), "DateObj": d, "Jour": JOURS_FR[d.strftime("%A")],
        "Ligne 1": planning[d]["L1"], "Ligne 2": planning[d]["L2"],
        "Type": "FÉRIÉ/WE" if is_we_ferie(d) else "SEMAINE"
    } for d in jours])
                 
    return res_df, sc, sc_we, n_l1, n_l2

# ==========================================
# 6. GÉNÉRATION EXCEL (A3 PAYSAGE)
# ==========================================
def generer_excel_liste(df_planning, d_sc, d_sc_we, d_nbl1, d_nbl2):
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    f_title = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    font_title = Font(color="FFFFFF", bold=True, size=18) 
    f_head = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    font_head = Font(color="FFFFFF", bold=True, size=14) 
    font_data = Font(size=13) 
    
    b_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    f_we = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    f_sep = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    df_planning['Annee'] = df_planning['DateObj'].dt.year
    df_planning['MoisNum'] = df_planning['DateObj'].dt.month

    ws_l1 = wb.create_sheet(title="Ligne 1")
    ws_l2 = wb.create_sheet(title="Ligne 2")

    for ws in [ws_l1, ws_l2]:
        ws.page_setup.paperSize = ws.PAPERSIZE_A3
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = False 
        ws.page_setup.fitToWidth = 1

    month_idx = 0
    for (annee, mois), group in df_planning.groupby(['Annee', 'MoisNum']):
        titre_mois = f"{MOIS_FR[mois].upper()} {annee}"
        col_offset = month_idx * 5 + 1 
        
        for ws in [ws_l1, ws_l2]:
            ws.column_dimensions[get_column_letter(col_offset)].width = 12      
            ws.column_dimensions[get_column_letter(col_offset + 1)].width = 14  
            ws.column_dimensions[get_column_letter(col_offset + 2)].width = 25  
            ws.column_dimensions[get_column_letter(col_offset + 3)].width = 24  
            
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 25
            
            ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset+3)
            c_title = ws.cell(row=1, column=col_offset, value=titre_mois)
            c_title.fill, c_title.font, c_title.alignment, c_title.border = f_title, font_title, align_center, b_thin
            for c_b in range(col_offset + 1, col_offset + 4): ws.cell(row=1, column=c_b).border = b_thin
            
            en_tetes = ["Jour", "Date", "Astreinte Prévue", "Modification / Remplaçant"]
            for i, val in enumerate(en_tetes):
                c = ws.cell(row=2, column=col_offset + i, value=val)
                c.fill, c.font, c.border, c.alignment = f_head, font_head, b_thin, align_center

        row_offset = 3
        for _, row in group.iterrows():
            for ws, ligne in [(ws_l1, 'Ligne 1'), (ws_l2, 'Ligne 2')]:
                ws.row_dimensions[row_offset].height = 24 
                
                c1 = ws.cell(row=row_offset, column=col_offset, value=row['Jour'])
                c2 = ws.cell(row=row_offset, column=col_offset+1, value=row['Date'])
                c3 = ws.cell(row=row_offset, column=col_offset+2, value=row[ligne])
                c4 = ws.cell(row=row_offset, column=col_offset+3, value="")

                for c in [c1, c2, c3, c4]:
                    c.border = b_thin
                    c.font = font_data
                    c.alignment = align_center if c.column != col_offset+3 else align_left
                    if row['Type'] == "FÉRIÉ/WE": c.fill = f_we
                    
            row_offset += 1
            
        if month_idx > 0:
            sep_col = col_offset - 1
            for ws in [ws_l1, ws_l2]:
                ws.column_dimensions[get_column_letter(sep_col)].width = 2
                for r in range(1, row_offset): ws.cell(row=r, column=sep_col).fill = f_sep

        month_idx += 1

    ws_bilan = wb.create_sheet(title="Bilan Équité")
    ws_bilan.append(["Manipulateur", "Total Points", "Total Astreintes", "Total L1", "Total L2", "Nb Week-ends"])
    for idx, m in enumerate(d_sc.keys(), 2):
        donnees_m = [m, d_sc[m], d_nbl1[m] + d_nbl2[m], d_nbl1[m], d_nbl2[m], d_sc_we[m]]
        for col_idx, val in enumerate(donnees_m, 1): ws_bilan.cell(row=idx, column=col_idx, value=val).border = b_thin
    for col in range(1, 7):
        ws_bilan.column_dimensions[get_column_letter(col)].width = 20
        c_head = ws_bilan.cell(row=1, column=col)
        c_head.fill, c_head.font, c_head.border = f_head, font_head, b_thin

    wb.save(output)
    return output.getvalue()

# ==========================================
# 7. INTERFACE UTILISATEUR (UI)
# ==========================================
st.title("🏥 Planning de Radiologie Interventionnelle")

c_cfg, c_res = st.columns([1, 2.2])

with c_cfg:
    st.header("1. Période (Trimestre)")
    st.session_state.d_start = st.date_input("Début", datetime.now())
    st.session_state.d_end = st.date_input("Fin", datetime.now() + timedelta(days=90))

with st.sidebar:
    st.header("⚙️ GESTION DE L'ÉQUIPE")
    with st.expander("➕ Ajouter un manipulateur", expanded=False):
        n_name = st.text_input("Nom & Prénom")
        n_l1, n_l2 = st.checkbox("Fait la Ligne 1", True), st.checkbox("Fait la Ligne 2", True)
        if st.button("Ajouter à l'équipe") and n_name and n_name not in st.session_state.merms_data:
            st.session_state.merms_data[n_name] = {
                "lignes": ([1] if n_l1 else []) + ([2] if n_l2 else []),
                "score_cumule": 0, "score_we": 0, "nb_l1": 0, "nb_l2": 0, "pref_vendredi": False, "absences": [], "obl_l1": [], "obl_l2": []
            }
            sauvegarder_donnees(st.session_state.merms_data); st.rerun()

    st.write("---")
    st.write("**Équipe actuelle**")
    for m, data in list(st.session_state.merms_data.items()):
        c1, c2 = st.columns([4, 1])
        c1.caption(f"**{m}** ({'L1/L2' if len(data['lignes']) == 2 else f'L{data['lignes'][0]}'})")
        st.markdown('<div class="btn-supprimer">', unsafe_allow_html=True)
        if c2.button("🗑️", key=f"del_{m}"):
            del st.session_state.merms_data[m]
            sauvegarder_donnees(st.session_state.merms_data); st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

with c_cfg:
    st.write("---")
    st.header("2. Desiderata")
    for merm, data in st.session_state.merms_data.items():
        if st.button(f"👤 {merm} ({len(data['absences'])} j. posés)", key=f"btn_{merm}"):
            st.session_state.modal_ouvert = merm; st.rerun() 

if st.session_state.modal_ouvert: modal_desiderata(st.session_state.modal_ouvert)

with c_res:
    st.header("3. Génération & Export")
    st.markdown('<div class="btn-generer">', unsafe_allow_html=True)
    if st.button("🚀 CALCULER LA RÉPARTITION ÉQUITABLE", use_container_width=True):
        res = generer_planning(st.session_state.d_start, st.session_state.d_end)
        st.session_state.planning_final, st.session_state.scores_finaux, st.session_state.scores_we_finaux, st.session_state.nbl1_finaux, st.session_state.nbl2_finaux = res
    st.markdown('</div>', unsafe_allow_html=True)

    if 'planning_final' in st.session_state:
        st.write("---")
        excel_data = generer_excel_liste(st.session_state.planning_final, st.session_state.scores_finaux, st.session_state.scores_we_finaux, st.session_state.nbl1_finaux, st.session_state.nbl2_finaux)
        st.download_button(
            label="📥 TÉLÉCHARGEMENT EXCEL (MURAL A3)", data=excel_data,
            file_name=f"Planning_RI_{st.session_state.d_start.strftime('%m')}-{st.session_state.d_end.strftime('%m_%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True
        )
        st.write("---")
        
        t1, t2, t3 = st.tabs(["📋 PLANNING LIGNE 1", "📋 PLANNING LIGNE 2", "📈 BILAN D'ÉQUITÉ"])
        with t1: st.table(st.session_state.planning_final[["Jour", "Date", "Ligne 1", "Type"]])
        with t2: st.table(st.session_state.planning_final[["Jour", "Date", "Ligne 2", "Type"]])
        with t3:
            st.info("L'algorithme équilibre dans l'ordre : 1. Volontaires (Bonus) -> 2. Nbr Week-ends -> 3. Pénibilité (Points) -> 4. Total Astreintes -> 5. Ratio L1/L2.")
            st.table(pd.DataFrame({
                "Total Points (Charge)": st.session_state.scores_finaux,
                "Total Astreintes": {m: st.session_state.nbl1_finaux[m] + st.session_state.nbl2_finaux[m] for m in st.session_state.merms_data},
                "Total L1": st.session_state.nbl1_finaux, "Total L2": st.session_state.nbl2_finaux,
                "Nb Week-ends": st.session_state.scores_we_finaux
            }))
            
            st.markdown('<div class="btn-valider">', unsafe_allow_html=True)
            if st.button("💾 VALIDER CE TRIMESTRE ET SAUVEGARDER L'HISTORIQUE", use_container_width=True):
                for m in st.session_state.scores_finaux:
                    st.session_state.merms_data[m].update({
                        'score_cumule': st.session_state.scores_finaux[m], 'score_we': st.session_state.scores_we_finaux[m],
                        'nb_l1': st.session_state.nbl1_finaux[m], 'nb_l2': st.session_state.nbl2_finaux[m]
                    })
                sauvegarder_donnees(st.session_state.merms_data)
                st.success("✅ Historique sauvegardé ! La prochaine répartition corrigera les écarts restants.")
            st.markdown('</div>', unsafe_allow_html=True)
